import uuid
from pathlib import Path
from typing import Protocol, List, Optional, Dict

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from py_cover_letters.db.models import CoverLetter
from py_cover_letters.constants import COLUMN_MAPPING, SHEET_NAME
from py_cover_letters.enums import FilterType
from py_cover_letters.exceptions import CoverLetterException
from py_cover_letters.utils import backup_file


class CoverLetterManager(Protocol):

    def get(self, cover_letter_id: int) -> CoverLetter:
        ...

    def create(self, cover_letter: CoverLetter) -> CoverLetter:
        ...

    def delete(self, cover_letter: CoverLetter) -> bool:
        ...

    def update(self, project: CoverLetter) -> CoverLetter:
        ...

    def list(self) -> List[CoverLetter]:
        ...

    def filter(self, filter_type: FilterType) -> List[CoverLetter]:
        ...


class ExcelManager:
    def __init__(self, filename: Path, column_mapping: Optional[Dict[int, str]] = None,
                 sheet_name: str = SHEET_NAME, backup_folder: Optional[Path]= None):
        if backup_folder is None:
            raise Exception('xxxxx')
        else:
            self.backup_folder = backup_folder
        self.filename = filename
        if column_mapping is None:
            self.column_mapping = COLUMN_MAPPING
        else:
            self.column_mapping = column_mapping
        self.sheet_name = sheet_name
        self.columns = [col_name for _, col_name in self.column_mapping.items()]
        self.cover_letters: List[CoverLetter] = list()
        if not self.filename.exists():
            self.write_template()
        else:
            self.cover_letters = self.list()

    def write_template(self):
        if self.filename.exists():
            error_msg = f'Cannot overwrite template {self.filename}'
            raise CoverLetterException(error_msg)

        wb = Workbook()
        sheet = wb.create_sheet(self.sheet_name, 0)
        row = 1
        col = 1
        for column_name in self.columns:
            sheet.cell(row=row, column=col, value=column_name)
            col += 1
        row += 1

        wb.save(self.filename)

    def list(self) -> List[CoverLetter]:
        cover_letters = list()
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        last_row = sheet.max_row + 1
        for row in range(2, last_row):
            cover_letter_dict = dict()
            for col, name in self.column_mapping.items():
                cell_obj = sheet.cell(row=row, column=col)
                value = cell_obj.value
                cover_letter_dict[name] = value
            try:
                cover_letter = CoverLetter(**cover_letter_dict)
                cover_letters.append(cover_letter)
            except Exception as e:
                error_message = f'Unexpected error on row {row}. Type: {e.__class__.__name__} Error: {e}'
                raise CoverLetterException(error_message)
        return cover_letters

    def update_not_saved(self, commit=True) -> List[CoverLetter]:
        not_saved = [x for x in self.cover_letters if x.id is None]
        for cover_letter in not_saved:
            cover_letter.id = uuid.uuid4().int
            print(cover_letter)
        if commit:
            backup_file(self.filename, self.backup_folder)
            self.filename.unlink()
            
        return not_saved
