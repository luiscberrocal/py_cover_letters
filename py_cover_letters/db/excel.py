from pathlib import Path
from typing import Dict, Any, List

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from .managers import CoverLetterManager
from .models import CoverLetter
from ..exceptions import CoverLetterException


def simple_write_to_excel(filename: Path, headers: Dict[str, Any], lines: List[Dict[str, Any]]):
    """Writes to excel a List of dictionaries. The headers keys must match the keys of the values to write.
    The headers must contain a dictionary with the title key."""
    wb = Workbook()
    sheet = wb.create_sheet()
    row = 1
    col = 1
    for key, header in headers.items():
        sheet.cell(row=row, column=col, value=header['title'])
        col += 1
    row += 1
    for line in lines:
        col = 1
        for key in headers.keys():
            value = line.get(key)
            if value is not None:
                sheet.cell(row=row, column=col, value=value)
            col += 1
        row += 1

    wb.save(filename)


COLUMN_MAPPING = {
    1: 'id',
    2: 'company_name',
    3: 'position_name',
    4: 'greeting',
    5: 'to_email',
    6: 'cover_template',
    7: 'date_sent_via_email',
    8: 'date_generated'
}


class ExcelCoverLetterManager:

    def __init__(self, filename: Path, column_mapping: Dict[int, str],
                 sheet_name: str = 'Cover letters'):
        self.filename = filename
        self.column_mapping = column_mapping
        self.columns = [col_name for _, col_name in self.column_mapping.items()]
        self.sheet_name = sheet_name

    def write_template(self):
        if self.filename.exists():
            error_msg = f'Cannot overwrite template {self.filename}'
            raise CoverLetterException(error_msg)

        wb = Workbook()
        sheet = wb.create_sheet(self.sheet_name, 0)
        row = 1
        col = 1
        for column_name in self.columns:
            sheet.cell(row=row, column=col, value=column_name)
            col += 1
        row += 1

        wb.save(self.filename)

    def read(self) -> List[CoverLetter]:
        cover_letters = list()
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        last_row = sheet.max_row + 1
        for row in range(2, last_row):
            cover_letter_dict = dict()
            for col, name in self.column_mapping.items():
                cell_obj = sheet.cell(row=row, column=col)
                value = cell_obj.value
                cover_letter_dict[name] = value
            try:
                cover_letter = CoverLetter(**cover_letter_dict)
                cover_letters.append(cover_letter)
            except Exception as e:
                error_message = f'Unexpected error on row {row}. Type: {e.__class__.__name__} Error: {e}'
                raise CoverLetterException(error_message)
        return cover_letters

    def add(self, cover_letters: List[CoverLetter]):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheet_name]
        row = sheet.max_row + 1
        for cover_letter in cover_letters:
            for col, attribute_name in self.column_mapping.items():
                value = getattr(cover_letter, attribute_name)
                sheet.cell(row=row, column=col, value=value)
            row += 1
        wb.save(self.filename)
